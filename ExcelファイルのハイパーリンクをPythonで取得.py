import openpyxl

'''
参考
#https://www.yuanfudao.com/tutor-ybc-course-api/pymodule/ybc_module/o_openpyxl.html
'''

#Sheetを開く
wb = openpyxl.load_workbook('C:/Users/tong/Desktop/他商材について.xlsx')
ws = wb['法人ページ']

for i in range(1,ws.max_row+1):
    #行数を取得
    
    cl=ws.cell(row=i, column=2)
    #セルオブジェクトを取得
    
    try:
        url=cl.hyperlink.target
        #セルのハイパーリンクの内容を取得
    
        v=cl.value
        #セルの値を取得
        
        print(v.ljust(100),url)
    
    except:
        pass
    
